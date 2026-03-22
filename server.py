#!/usr/bin/env python3
"""
JoshOS Mobile — Flask Backend
PIN auth, JWT sessions, rate limiting, all API routes.
Serves React PWA from /dist.
"""

import os
import sys
import json
import secrets
import hashlib
import subprocess
import threading
import time
import uuid
import logging
from datetime import datetime, timedelta
from pathlib import Path
from functools import wraps

from flask import Flask, request, jsonify, send_from_directory
import jwt

# ── Validate secrets on startup ──────────────────────────────────────────────

REQUIRED_SECRETS = ["JOSHOS_PIN", "ANTHROPIC_API_KEY"]
missing = [s for s in REQUIRED_SECRETS if not os.environ.get(s)]
if missing:
    print(f"FATAL: Missing required secrets: {', '.join(missing)}", file=sys.stderr)
    print("Set them in Replit Secrets or environment variables.", file=sys.stderr)
    sys.exit(1)

JWT_SECRET = os.environ.get("JWT_SECRET_KEY") or secrets.token_hex(32)
JOSHOS_PIN = os.environ["JOSHOS_PIN"]
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]

# ── App setup ────────────────────────────────────────────────────────────────

app = Flask(__name__, static_folder="dist", static_url_path="")

DATA_DIR = Path(__file__).parent / "data"
BRIEFINGS_DIR = DATA_DIR / "briefings"
CONTACTS_DIR = DATA_DIR / "contacts"

for d in [BRIEFINGS_DIR, CONTACTS_DIR]:
    d.mkdir(parents=True, exist_ok=True)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("joshos-mobile")

# ── Rate limiting (simple in-memory) ─────────────────────────────────────────

_login_attempts = {}  # ip -> {"count": int, "first_at": float, "locked_until": float}

def check_rate_limit(ip):
    """Returns (allowed: bool, retry_after: int or None)."""
    now = time.time()
    rec = _login_attempts.get(ip, {"count": 0, "first_at": now, "locked_until": 0})

    if now < rec.get("locked_until", 0):
        remaining = int(rec["locked_until"] - now)
        return False, remaining

    # Reset window if > 10 minutes since first attempt
    if now - rec.get("first_at", now) > 600:
        rec = {"count": 0, "first_at": now, "locked_until": 0}

    return True, None

def record_failed_attempt(ip):
    now = time.time()
    rec = _login_attempts.get(ip, {"count": 0, "first_at": now, "locked_until": 0})

    if now - rec.get("first_at", now) > 600:
        rec = {"count": 0, "first_at": now, "locked_until": 0}

    rec["count"] = rec.get("count", 0) + 1

    if rec["count"] >= 5:
        rec["locked_until"] = now + 1800  # 30 minute lockout
        rec["count"] = 0
        rec["first_at"] = now

    _login_attempts[ip] = rec
    return rec["count"]

def reset_attempts(ip):
    _login_attempts.pop(ip, None)

# ── JWT helpers ──────────────────────────────────────────────────────────────

def create_token():
    payload = {
        "sub": "josh",
        "iat": datetime.utcnow(),
        "exp": datetime.utcnow() + timedelta(hours=12),
        "jti": str(uuid.uuid4()),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")

def verify_token(token):
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None

def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth_header = request.headers.get("Authorization", "")
        if not auth_header.startswith("Bearer "):
            return jsonify({"error": "Authentication required", "code": 401}), 401
        token = auth_header[7:]
        payload = verify_token(token)
        if not payload:
            return jsonify({"error": "Invalid or expired token", "code": 401}), 401
        return f(*args, **kwargs)
    return decorated

# ── Background jobs ──────────────────────────────────────────────────────────

_jobs = {}  # job_id -> {"status": str, "output": str, "started": float}

def run_job(job_id, script_name):
    """Run a script in the background and track its output."""
    _jobs[job_id]["status"] = "running"
    try:
        result = subprocess.run(
            ["python3", "-c", f"print('Script {script_name} would run here — connect to iMac scripts for live execution')"],
            capture_output=True, text=True, timeout=300,
        )
        _jobs[job_id]["output"] = result.stdout + result.stderr
        _jobs[job_id]["status"] = "done" if result.returncode == 0 else "error"
    except Exception as e:
        _jobs[job_id]["output"] = str(e)
        _jobs[job_id]["status"] = "error"

# ── Auth routes (no @require_auth) ───────────────────────────────────────────

@app.route("/auth/login", methods=["POST"])
def login():
    ip = request.remote_addr
    allowed, retry_after = check_rate_limit(ip)

    if not allowed:
        return jsonify({
            "error": f"Too many attempts. Try again in {retry_after // 60} minutes.",
            "code": 429, "locked": True, "retry_after": retry_after,
        }), 429

    body = request.get_json() or {}
    pin = body.get("pin", "")

    if not pin or pin != JOSHOS_PIN:
        remaining = record_failed_attempt(ip)
        attempts_left = max(0, 5 - remaining)
        return jsonify({
            "error": "Invalid PIN",
            "code": 403,
            "attempts_left": attempts_left,
        }), 403

    reset_attempts(ip)
    token = create_token()
    return jsonify({"token": token, "expires_in": 43200})

# ── API routes ───────────────────────────────────────────────────────────────

@app.route("/api/today", methods=["GET"])
@require_auth
def api_today():
    today = datetime.now().strftime("%Y-%m-%d")
    morning = None
    evening = None

    for f in sorted(BRIEFINGS_DIR.glob("*.md"), reverse=True):
        if today in f.name:
            content = f.read_text()
            if "morning" in f.name.lower() or "daily" in f.name.lower():
                morning = {"filename": f.name, "content": content, "type": "morning"}
            elif "evening" in f.name.lower() or "intel" in f.name.lower():
                evening = {"filename": f.name, "content": content, "type": "evening"}

    return jsonify({
        "date": today,
        "day": datetime.now().strftime("%A"),
        "morning": morning,
        "evening": evening,
    })

@app.route("/api/briefings", methods=["GET"])
@require_auth
def api_briefings():
    files = sorted(BRIEFINGS_DIR.glob("*.md"), reverse=True)
    briefings = []
    for f in files:
        content = f.read_text()
        btype = "morning"
        if "evening" in f.name.lower() or "intel" in f.name.lower():
            btype = "evening"
        elif "capabilities" in f.name.lower():
            btype = "capabilities"
        elif "roadmap" in f.name.lower():
            btype = "roadmap"

        preview = content[:200].replace("#", "").strip()
        briefings.append({
            "id": f.stem,
            "filename": f.name,
            "type": btype,
            "preview": preview,
            "date": f.stem.split("-")[-3:] if len(f.stem.split("-")) >= 3 else [],
        })
    return jsonify({"briefings": briefings, "count": len(briefings)})

@app.route("/api/briefings/<bid>", methods=["GET"])
@require_auth
def api_briefing_detail(bid):
    for f in BRIEFINGS_DIR.glob("*.md"):
        if f.stem == bid:
            return jsonify({"id": bid, "filename": f.name, "content": f.read_text()})
    return jsonify({"error": "Briefing not found", "code": 404}), 404

@app.route("/api/contacts", methods=["GET"])
@require_auth
def api_contacts():
    xlsx_files = list(CONTACTS_DIR.glob("*.xlsx"))
    if not xlsx_files:
        return jsonify({"contacts": [], "count": 0})

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_files[0], read_only=True)
        ws = wb.active
        contacts = []
        for row in range(2, ws.max_row + 1):
            first = ws.cell(row=row, column=1).value
            if not first:
                continue
            contacts.append({
                "id": row,
                "first_name": str(first or ""),
                "last_name": str(ws.cell(row=row, column=2).value or ""),
                "title": str(ws.cell(row=row, column=3).value or ""),
                "company": str(ws.cell(row=row, column=4).value or ""),
                "account": str(ws.cell(row=row, column=5).value or ""),
                "linkedin": str(ws.cell(row=row, column=6).value or ""),
                "status": str(ws.cell(row=row, column=7).value or ""),
                "dossier_link": str(ws.cell(row=row, column=8).value or ""),
                "notes": str(ws.cell(row=row, column=9).value or ""),
                "date_generated": str(ws.cell(row=row, column=10).value or ""),
            })
        wb.close()
        return jsonify({"contacts": contacts, "count": len(contacts)})
    except Exception as e:
        return jsonify({"error": str(e), "code": 500}), 500

@app.route("/api/contacts/<int:cid>", methods=["GET"])
@require_auth
def api_contact_detail(cid):
    xlsx_files = list(CONTACTS_DIR.glob("*.xlsx"))
    if not xlsx_files:
        return jsonify({"error": "No contacts file", "code": 404}), 404

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_files[0], read_only=True)
        ws = wb.active
        if cid < 2 or cid > ws.max_row:
            return jsonify({"error": "Contact not found", "code": 404}), 404

        contact = {
            "id": cid,
            "first_name": str(ws.cell(row=cid, column=1).value or ""),
            "last_name": str(ws.cell(row=cid, column=2).value or ""),
            "title": str(ws.cell(row=cid, column=3).value or ""),
            "company": str(ws.cell(row=cid, column=4).value or ""),
            "account": str(ws.cell(row=cid, column=5).value or ""),
            "linkedin": str(ws.cell(row=cid, column=6).value or ""),
            "status": str(ws.cell(row=cid, column=7).value or ""),
            "dossier_link": str(ws.cell(row=cid, column=8).value or ""),
            "notes": str(ws.cell(row=cid, column=9).value or ""),
            "date_generated": str(ws.cell(row=cid, column=10).value or ""),
        }
        wb.close()
        return jsonify(contact)
    except Exception as e:
        return jsonify({"error": str(e), "code": 500}), 500

@app.route("/api/contacts", methods=["POST"])
@require_auth
def api_add_contact():
    body = request.get_json() or {}
    if not body.get("confirmed"):
        return jsonify({"error": "Confirmation required", "code": 400}), 400

    return jsonify({"message": "Contact added — sync to iMac EMS to trigger research", "status": "queued"})

@app.route("/api/ask", methods=["POST"])
@require_auth
def api_ask():
    body = request.get_json() or {}
    question = body.get("question", "").strip()
    if not question:
        return jsonify({"error": "No question provided", "code": 400}), 400

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        resp = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            system="You are JoshOS, Josh Larivee's executive AI assistant. Josh is Director of Life Sciences at AWS managing pharma accounts (Pfizer, BMS, Lilly, Novartis, Cigna, Elevance). He also owns Three Rivers Slab Co. and MadSprings Cookies. Answer in concise prose paragraphs. Be specific and actionable.",
            messages=[{"role": "user", "content": question}],
        )
        answer = "".join(b.text for b in resp.content if hasattr(b, "text"))
        return jsonify({"question": question, "answer": answer})
    except Exception as e:
        return jsonify({"error": str(e), "code": 500}), 500

@app.route("/api/actions/run", methods=["POST"])
@require_auth
def api_run_action():
    body = request.get_json() or {}
    if not body.get("confirmed"):
        return jsonify({"error": "Confirmation required", "code": 400}), 400

    script = body.get("script", "")
    valid = {"briefing", "evening", "capabilities", "roadmap", "surface"}
    if script not in valid:
        return jsonify({"error": f"Invalid script. Valid: {', '.join(valid)}", "code": 400}), 400

    job_id = str(uuid.uuid4())[:8]
    _jobs[job_id] = {"status": "queued", "output": "", "started": time.time(), "script": script}
    thread = threading.Thread(target=run_job, args=(job_id, script), daemon=True)
    thread.start()

    return jsonify({"job_id": job_id, "status": "queued"})

@app.route("/api/actions/<job_id>", methods=["GET"])
@require_auth
def api_job_status(job_id):
    job = _jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found", "code": 404}), 404
    return jsonify({
        "job_id": job_id,
        "status": job["status"],
        "output": job["output"][-2000:] if job["output"] else "",
    })

@app.route("/api/upload", methods=["POST"])
@require_auth
def api_upload():
    """Upload briefing or contact files from iMac sync script."""
    body = request.get_json() or {}
    filename = body.get("filename", "")
    content = body.get("content", "")
    folder = body.get("folder", "briefings")

    if not filename or not content:
        return jsonify({"error": "filename and content required", "code": 400}), 400

    target_dir = BRIEFINGS_DIR if folder == "briefings" else CONTACTS_DIR
    target = target_dir / filename
    target.write_text(content)
    return jsonify({"message": f"Uploaded {filename}", "path": str(target)})

# ── Life OS Routes ───────────────────────────────────────────────────────────

LIFE_DATA_FILE = DATA_DIR / "life_data.json"

@app.route("/api/life", methods=["GET"])
@require_auth
def api_life():
    if LIFE_DATA_FILE.exists():
        return jsonify(json.loads(LIFE_DATA_FILE.read_text()))
    return jsonify({
        "readiness": {}, "health_pulse": "", "wealth_pulse": "",
        "thesis_verdicts": [], "portfolio_summary": {},
        "last_updated": None, "status": "no data synced yet",
    })

@app.route("/api/life/sync", methods=["POST"])
@require_auth
def api_life_sync():
    body = request.get_json() or {}
    body["synced_at"] = datetime.now().isoformat()
    LIFE_DATA_FILE.write_text(json.dumps(body, indent=2, default=str))
    logger.info("Life data synced from iMac")
    return jsonify({"message": "Life data synced", "synced_at": body["synced_at"]})

@app.route("/api/life/restaurant", methods=["POST"])
@require_auth
def api_life_restaurant():
    body = request.get_json() or {}
    name = body.get("name", "")
    city = body.get("city", "")
    notes = body.get("notes", "")
    rating = body.get("rating", 0)
    if not name or not city:
        return jsonify({"error": "name and city required", "code": 400}), 400

    # Save locally on Replit
    rest_file = DATA_DIR / "restaurants.json"
    rests = json.loads(rest_file.read_text()) if rest_file.exists() else []
    entry = {"name": name, "city": city, "date_visited": datetime.now().strftime("%Y-%m-%d"),
             "notes": notes, "rating": rating}
    rests.append(entry)
    rest_file.write_text(json.dumps(rests, indent=2))
    return jsonify({"message": f"Logged {name} in {city}", "entry": entry})

@app.route("/api/life/thesis", methods=["GET"])
@require_auth
def api_life_thesis():
    if LIFE_DATA_FILE.exists():
        data = json.loads(LIFE_DATA_FILE.read_text())
        return jsonify({"theses": data.get("thesis_verdicts", [])})
    return jsonify({"theses": []})

@app.route("/api/life/food/<city>", methods=["GET"])
@require_auth
def api_life_food(city):
    guide_file = DATA_DIR / "food-guides" / f"{city.lower().replace(' ', '-')}.md"
    if guide_file.exists():
        return jsonify({"city": city, "guide": guide_file.read_text()})
    return jsonify({"city": city, "guide": None, "message": f"No guide cached for {city}. Generate from iMac."})

@app.route("/api/life/refresh", methods=["POST"])
@require_auth
def api_life_refresh():
    if LIFE_DATA_FILE.exists():
        data = json.loads(LIFE_DATA_FILE.read_text())
        data["cached"] = True
        return jsonify(data)
    return jsonify({"error": "No cached data. Run agents on iMac first.", "code": 404}), 404

# ── Serve React PWA ──────────────────────────────────────────────────────────

@app.route("/")
def serve_index():
    return send_from_directory(app.static_folder, "index.html")

@app.route("/<path:path>")
def serve_static(path):
    file_path = Path(app.static_folder) / path
    if file_path.exists():
        return send_from_directory(app.static_folder, path)
    return send_from_directory(app.static_folder, "index.html")

# ── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    logger.info(f"JoshOS Mobile starting on port {port}")
    app.run(host="0.0.0.0", port=port, debug=False)
